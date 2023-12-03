import copy
import logging
import random

import numpy as np
import torch
import wandb
import openpyxl
import time
from fedml import mlops
from fedml.ml.trainer.trainer_creator import create_model_trainer
from .client import Client, client_upload_time, client_aggregation_time
import matplotlib.pyplot as plt

global_client_num_in_total = 60
# global_client_num_per_round = 30

accuracy_list = []
loss_list = []
# 参数0
# 统计global_client_num_in_total个客户每个人的被选择次数
client_selected_times = [0 for i in range(global_client_num_in_total)]
plt.figure(1, figsize=(20, 10))

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 创建工作表
client_ws = wb.create_sheet('Clients Info')
# 写入损失指标的标头行
client_ws.append(['Round', 'ClientIdx', 'Loss', 'Accuracy', 'Time'])
# 创建工作表
round_ws = wb.create_sheet('Round Info')
# 写入精度指标的标头行
round_ws.append(['Round', 'Loss', 'Accuracy', 'Time', 'Selected Client Indexs', 'Total Selected Client Times'])
# 设置时间间隔（以秒为单位）
interval = 5


class FedCSAPI(object):
    def __init__(self, args, device, dataset, model):
        self.device = device
        self.args = args

        [
            train_data_num,
            test_data_num,
            train_data_global,
            test_data_global,
            train_data_local_num_dict,
            train_data_local_dict,
            test_data_local_dict,
            class_num,
        ] = dataset

        self.train_global = train_data_global
        self.test_global = test_data_global
        self.val_global = None
        self.train_data_num_in_total = train_data_num
        self.test_data_num_in_total = test_data_num
        self.args.client_num_in_total = global_client_num_in_total  # added
        # self.args.client_num_per_round = global_client_num_per_round  # added
        self.client_list = []
        self.client_train_prob = []  # 客户训练概成功率列表
        self.train_data_local_num_dict = train_data_local_num_dict
        self.train_data_local_dict = train_data_local_dict
        self.test_data_local_dict = test_data_local_dict

        logging.info("model = {}".format(model))
        self.model_trainer = create_model_trainer(model, args)
        self.model = model
        logging.info("self.model_trainer = {}".format(self.model_trainer))

        self._setup_clients(
            train_data_local_num_dict, train_data_local_dict, test_data_local_dict, self.model_trainer,
        )

    def _setup_clients(
        self, train_data_local_num_dict, train_data_local_dict, test_data_local_dict, model_trainer,
    ):
        logging.info("############setup_clients (START)#############")
        for client_idx in range(self.args.client_num_in_total):
            c = Client(
                client_idx,
                train_data_local_dict[client_idx],
                test_data_local_dict[client_idx],
                train_data_local_num_dict[client_idx],
                self.args,
                self.device,
                model_trainer,
            )
            self.client_list.append(c)
            self.client_train_prob.append(0.5)
        logging.info("############setup_clients (END)#############")

    def train(self):
        logging.info("self.model_trainer = {}".format(self.model_trainer))
        w_global = self.model_trainer.get_model_params()
        mlops.log_training_status(mlops.ClientConstants.MSG_MLOPS_CLIENT_STATUS_TRAINING)
        mlops.log_aggregation_status(mlops.ServerConstants.MSG_MLOPS_SERVER_STATUS_RUNNING)
        mlops.log_round_info(self.args.comm_round, -1)
        for round_idx in range(self.args.comm_round):

            logging.info("################Communication round : {}".format(round_idx))

            w_locals = []

            """
            for scalability: following the original FedAvg algorithm, we uniformly sample a fraction of clients in each round.
            Instead of changing the 'Client' instances, our implementation keeps the 'Client' instances and then updates their local dataset 
            """
            client_indexes = self._client_sampling(
                round_idx, self.args.client_num_in_total, self.args.client_num_per_round
            )
            logging.info("client_indexes = " + str(client_indexes))

            for idx, client in enumerate(self.client_list):
                # update dataset
                # 判断如果idx是client_indexes中的某一client的下标，那么就更新这个client的数据集
                if client.client_idx in client_indexes:
                    client_idx = client.client_idx
                    client.update_local_dataset(
                        client_idx,
                        self.train_data_local_dict[client_idx],
                        self.test_data_local_dict[client_idx],
                        self.train_data_local_num_dict[client_idx],
                    )

                    # train on new dataset
                    mlops.event("train", event_started=True,event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
                    w = client.train(copy.deepcopy(w_global))
                    mlops.event("train", event_started=False,event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
                    # self.logging.info("local weights = " + str(w))
                    if self.judge_model(self.client_train_prob[client.client_idx]) == 1: # 判断是否成功返回模型
                        w_locals.append((client.get_sample_number(), copy.deepcopy(w)))

            # 借助client_selected_times统计global_client_num_in_total个客户每个人的被选择次数
            for i in client_indexes:
                client_selected_times[i] += 1

            # update global weights
            mlops.event("agg", event_started=True, event_value=str(round_idx))
            w_global = self._aggregate(w_locals)

            self.model_trainer.set_model_params(w_global)
            mlops.event("agg", event_started=False, event_value=str(round_idx))

            # test results
            # at last round
            train_acc, train_loss, test_acc, test_loss = 0, 0, 0, 0
            if round_idx == self.args.comm_round - 1:
                train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)
            # per {frequency_of_the_test} round
            elif round_idx % self.args.frequency_of_the_test == 0:
                if self.args.dataset.startswith("stackoverflow"):
                    self._local_test_on_validation_set(round_idx)
                else:
                    train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)

            mlops.log_round_info(self.args.comm_round, round_idx)

            round_ws.append([round_idx,
                             train_loss,
                             train_acc,
                             time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),
                             str(client_indexes),
                             str(client_selected_times)])

            # 保存Excel文件到self.args.excel_save_path+文件名
            wb.save(self.args.excel_save_path + self.args.model + "_[" + self.args.dataset
                    + "]_fedcs_training_results_NIID" + str(self.args.experiment_niid_level) + ".xlsx")
            # 休眠一段时间，以便下一个循环开始前有一些时间
            time.sleep(interval)

        mlops.log_training_finished_status()
        mlops.log_aggregation_finished_status()

    def judge_model(self,prob): # 基于随机数结合概率判断是否成功返回模型
        random_number = random.random()  # 生成0到1之间的随机数
        if random_number < prob:
            return 1  # 成功返回模型
        else:
            return 0

    # def _client_sampling(self, round_idx, client_num_in_total, client_num_per_round):
    #     if client_num_in_total == client_num_per_round:
    #         client_indexes = [client_index for client_index in range(client_num_in_total)]
    #     else:
    #         num_clients = min(client_num_per_round, client_num_in_total)
    #         np.random.seed(round_idx)  # make sure for each comparison, we are selecting the same clients each round
    #         client_indexes = np.random.choice(range(client_num_in_total), num_clients, replace=False)
    #     logging.info("client_indexes = %s" % str(client_indexes))
    #     return client_indexes

    def _client_sampling(self, round_idx, client_num_in_total, client_num_per_round):
        # Calculate the estimated time elapsed for each client
        Tcs = 1  # Client Selection step time
        Td = 0  # Total time for clients in S
        Theta = 0  # Current time
        Tround = 10  # Maximum time for each round
        K0 = list(range(client_num_in_total))  # Set of randomly selected clients
        client_indexes = []  # Selected clients
        t = Tcs  # Current time
        while len(K0) > client_num_in_total - client_num_per_round and t < Tround:
            x = None
            max_val = float('-inf')
            for k in K0:
                Td_S_k = Td + self.client_list[k].estimate_upload_time(round_idx)  # Estimated time for k to upload its model
                Theta_S_k = Theta + Td_S_k  # Estimated time for k to complete the whole process
                val = Tcs + Td_S_k + max(0, Theta_S_k - Theta)  # Estimated time for k to join S
                if val > max_val:
                    max_val = val
                    x = k
            if x is None:
                break
            Td += self.client_list[x].estimate_upload_time(round_idx)
            Theta0 = Theta + self.client_list[x].estimate_upload_time(round_idx)
            t = Tcs + Td + max(0, Theta0 - Theta) + self.client_list[x].estimate_aggregation_time(round_idx)
            if t < Tround:
                client_indexes.append(x)
                K0.remove(x)
                Theta = Theta0
        logging.info("Selected clients: %s" % str(client_indexes))
        return client_indexes

    def _generate_validation_set(self, num_samples=10000):
        test_data_num = len(self.test_global.dataset)
        sample_indices = random.sample(range(test_data_num), min(num_samples, test_data_num))
        subset = torch.utils.data.Subset(self.test_global.dataset, sample_indices)
        sample_testset = torch.utils.data.DataLoader(subset, batch_size=self.args.batch_size)
        self.val_global = sample_testset

    def _aggregate(self, w_locals):
        training_num = 0
        for idx in range(len(w_locals)):
            (sample_num, averaged_params) = w_locals[idx]
            training_num += sample_num

        (sample_num, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            print("______k = " + str(k))
            for i in range(0, len(w_locals)):
                local_sample_number, local_model_params = w_locals[i]
                w = local_sample_number / training_num
                if i == 0:
                    averaged_params[k] = local_model_params[k] * w
                else:
                    averaged_params[k] += local_model_params[k] * w
        return averaged_params

    def _aggregate_noniid_avg(self, w_locals):
        """
        The old aggregate method will impact the model performance when it comes to Non-IID setting
        Args:
            w_locals:
        Returns:
        """
        (_, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            temp_w = []
            for (_, local_w) in w_locals:
                temp_w.append(local_w[k])
            averaged_params[k] = sum(temp_w) / len(temp_w)
        return averaged_params

    def _local_test_on_all_clients(self, round_idx):

        logging.info("################local_test_on_all_clients : {}".format(round_idx))

        train_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        test_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        client = self.client_list[0]

        for client_idx in range(self.args.client_num_in_total):
            """
            Note: for datasets like "fed_CIFAR100" and "fed_shakespheare",
            the training client number is larger than the testing client number
            """
            if self.test_data_local_dict[client_idx] is None:
                continue
            client.update_local_dataset(
                0,
                self.train_data_local_dict[client_idx],
                self.test_data_local_dict[client_idx],
                self.train_data_local_num_dict[client_idx],
            )
            # train data
            train_local_metrics = client.local_test(False)
            train_metrics["num_samples"].append(copy.deepcopy(train_local_metrics["test_total"]))
            train_metrics["num_correct"].append(copy.deepcopy(train_local_metrics["test_correct"]))
            train_metrics["losses"].append(copy.deepcopy(train_local_metrics["test_loss"]))

            # test data
            test_local_metrics = client.local_test(True)
            test_metrics["num_samples"].append(copy.deepcopy(test_local_metrics["test_total"]))
            test_metrics["num_correct"].append(copy.deepcopy(test_local_metrics["test_correct"]))
            test_metrics["losses"].append(copy.deepcopy(test_local_metrics["test_loss"]))

            client_ws.append([round_idx,
                              client_idx,
                              train_metrics["losses"][client_idx] / train_metrics["num_samples"][client_idx],
                              train_metrics["num_correct"][client_idx] / train_metrics["num_samples"][client_idx],
                              time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())])

        # test on training dataset
        train_acc = sum(train_metrics["num_correct"]) / sum(train_metrics["num_samples"])
        train_loss = sum(train_metrics["losses"]) / sum(train_metrics["num_samples"])

        # test on test dataset
        test_acc = sum(test_metrics["num_correct"]) / sum(test_metrics["num_samples"])
        test_loss = sum(test_metrics["losses"]) / sum(test_metrics["num_samples"])

        stats = {"training_acc": train_acc, "training_loss": train_loss}
        if self.args.enable_wandb:
            wandb.log({"Train/Acc": train_acc, "round": round_idx})
            wandb.log({"Train/Loss": train_loss, "round": round_idx})

        # 绘制精度图
        accuracy_list.append(train_acc)
        loss_list.append(train_loss)
        plot_accuracy_and_loss(self, round_idx)

        mlops.log({"Train/Acc": train_acc, "round": round_idx})
        mlops.log({"Train/Loss": train_loss, "round": round_idx})
        logging.info(stats)

        stats = {"test_acc": test_acc, "test_loss": test_loss}
        if self.args.enable_wandb:
            wandb.log({"Test/Acc": test_acc, "round": round_idx})
            wandb.log({"Test/Loss": test_loss, "round": round_idx})

        mlops.log({"Test/Acc": test_acc, "round": round_idx})
        mlops.log({"Test/Loss": test_loss, "round": round_idx})
        logging.info(stats)

        return train_acc, train_loss, test_acc, test_loss

    def _local_test_on_validation_set(self, round_idx):

        logging.info("################local_test_on_validation_set : {}".format(round_idx))

        if self.val_global is None:
            self._generate_validation_set()

        client = self.client_list[0]
        client.update_local_dataset(0, None, self.val_global, None)
        # test data
        test_metrics = client.local_test(True)

        if self.args.dataset == "stackoverflow_nwp":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {"test_acc": test_acc, "test_loss": test_loss}
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})

        elif self.args.dataset == "stackoverflow_lr":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_pre = test_metrics["test_precision"] / test_metrics["test_total"]
            test_rec = test_metrics["test_recall"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {
                "test_acc": test_acc,
                "test_pre": test_pre,
                "test_rec": test_rec,
                "test_loss": test_loss,
            }
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Pre": test_pre, "round": round_idx})
                wandb.log({"Test/Rec": test_rec, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Pre": test_pre, "round": round_idx})
            mlops.log({"Test/Rec": test_rec, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})
        else:
            raise Exception("Unknown format to log metrics for dataset {}!" % self.args.dataset)

        logging.info(stats)


# 定义绘制精度图的函数
def plot_accuracy_and_loss(self, round_idx):
    plt.ion()

    print("accuracy_list: ", accuracy_list)
    print("loss_list: ", loss_list)

    plt.clf()
    plt.suptitle("FedCS" + "_[" + self.args.dataset + "]_NIID" + str(self.args.experiment_niid_level))

    # 第1个子图
    plt.subplot(2, 3, 1)
    plt.title("accuracy")
    plt.xlabel("num of epoch")
    plt.ylabel("value of accuracy")
    plt.plot(range(1, len(accuracy_list) + 1), accuracy_list, 'b-', linewidth=2)

    # 第2个子图
    plt.subplot(2, 3, 2)
    plt.title("loss")
    plt.xlabel("num of epoch")
    plt.ylabel("value of loss")
    plt.plot(range(1, len(loss_list) + 1), loss_list, 'b-', linewidth=2)

    # 第3个子图，使用条形图展示每个客户的被选择次数
    plt.subplot(2, 3, 3)
    plt.title("num of selected")
    plt.xlabel("num of epoch")
    plt.ylabel("value of num of selected")
    plt.bar(range(1, len(client_selected_times) + 1), client_selected_times, width=0.5, fc='b')

    # 第4个子图，使用折线图展示每个客户的预估上传时间
    plt.subplot(2, 3, 4)
    plt.title("estimated upload time")
    plt.xlabel("Clients' Indexes")
    plt.ylabel("value of estimated upload time")
    for i in range(len(client_upload_time)):
        plt.plot(range(1, global_client_num_in_total + 1), client_upload_time[i], label="Round {}".format(i + 1))
    plt.legend()

    # 第5个子图，使用折线图展示每个客户的预估聚合时间
    plt.subplot(2, 3, 5)
    plt.title("estimated aggregation time")
    plt.xlabel("Clients' Indexes")
    plt.ylabel("value of estimated aggregation time")
    for i in range(len(client_aggregation_time)):
        plt.plot(range(1, global_client_num_in_total + 1), client_aggregation_time[i], label="Round {}".format(i + 1))
    plt.legend()

    plt.tight_layout()
    plt.pause(0.03)
    plt.ioff()

    if (round_idx == self.args.comm_round - 1):
        plt.show()

    return
