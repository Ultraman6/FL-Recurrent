import copy
import logging
import random
from collections import OrderedDict
from math import log

import openpyxl
import time
import numpy as np
import torch
import wandb

from fedml import mlops
from fedml.ml.trainer.trainer_creator import create_model_trainer
from .client import Client
import matplotlib.pyplot as plt

from ..classical_vertical_fl.party_models import sigmoid

global_client_num_in_total = 60
global_client_num_per_round = 30

accuracy_list = []
loss_list = []
# 参数0
plt.figure(1, figsize=(10, 5))

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 创建工作表
client_ws = wb.create_sheet('Clients Info')
# 写入损失指标的标头行
client_ws.append(['Round', 'ClientIdx', 'Loss', 'Accuracy', 'Time'])
# 创建工作表
round_ws = wb.create_sheet('Round Info')
# 写入精度指标的标头行
round_ws.append(['Round', 'Loss', 'Accuracy', 'Time', 'Selected Client Indexs', 'Total Selected Client Times'])
# 设置时间间隔（以秒为单位）
interval = 5


class FedABAPI(object):
    def __init__(self, args, device, dataset, model):
        self.device = device
        self.args = args
        [
            train_data_num,
            test_data_num,
            train_data_global,
            test_data_global,
            train_data_local_num_dict,
            train_data_local_dict,
            test_data_local_dict,
            class_num,
            # Q_bounds,  # 质量属性值的范畴 [qmin，qmax]
            # Q_weights,  # 质量属性权重 [w1,w2,...,wn]
            total_budget
        ] = dataset

        self.train_global = train_data_global
        self.test_global = test_data_global
        self.val_global = None
        self.train_data_num_in_total = train_data_num
        self.test_data_num_in_total = test_data_num
        self.Q_weights = [0.4,0.4,0.2]   # 读入BS规定的客户非价格质量属性权重 w
        # FedAB参数
        self.args.client_num_in_total = global_client_num_in_total  # 总客户数
        self.client_list = []  # 总客户集合 N
        self.client_indexes = []  # 每轮选择客户集 W
        self.global_client_num_in_total = global_client_num_in_total  # 总客户数 N
        self.args.client_num_per_round = global_client_num_per_round  # 每轮需选中客户数 K
        self.client_train_prob = []  # 客户训练概成功率φ列表
        # self.client_train_noises = []  # 客户训练噪声系数ζ列表
        # 激励过程
        self.client_bids = []  # 客户报价列表
        self.client_qualities = []  # 客户质量列表
        self.client_scores = []  # 客户得分列表
        self.client_rewards = []  # 每轮次开始前清0，只更新选中客户的奖励
        self.client_avgRewards = []  # 基于历史信息的客户平均奖励
        self.client_UCBRewards = []  # 结合未来预测的客户UCB奖励
        self.client_UCB = []  # 客户UCB指标
        self.client_Payment = []  # 客户支付列表(全轮次)
        # 客户总被选中次数
        self.client_selected_times = [0 for i in range(self.global_client_num_in_total)]  # 统计客户被选择次数 β
        self.total_reward = 0  # 总奖励
        self.total_budget = total_budget  # 总预算
        # 客户训练数据参数
        self.train_data_local_num_dict = train_data_local_num_dict
        self.train_data_local_dict = train_data_local_dict
        self.test_data_local_dict = test_data_local_dict

        logging.info("model = {}".format(model))
        self.model_trainer = create_model_trainer(model, args)
        self.model = model
        logging.info("self.model_trainer = {}".format(self.model_trainer))

        self._setup_clients(
            train_data_local_num_dict, train_data_local_dict, test_data_local_dict, self.model_trainer,
        )

    def _setup_clients(
            self, train_data_local_num_dict, train_data_local_dict, test_data_local_dict, model_trainer,
    ):
        logging.info("############setup_clients (START)#############")
        for client_idx in range(self.args.client_num_in_total):
            c = Client(
                client_idx,
                train_data_local_dict[client_idx],
                test_data_local_dict[client_idx],
                train_data_local_num_dict[client_idx],
                self.args,
                self.device,
                model_trainer,
            )
            self.client_list.append(c)
            self.client_train_prob.append(1)  # 客户训练概成功率φ列表
            # self.client_train_noises.append(1)  # 客户训练噪声系数ζ列表
            self.client_bids.append(0)  # 客户报价列表
            self.client_qualities.append(0)  # 客户质量列表
            self.client_scores.append(0)  # 客户得分列表
            self.client_rewards.append(0)  # 每轮次开始清0，只更新选中客户的奖励
            self.client_avgRewards.append(0)  # 基于历史信息的客户平均奖励
            self.client_UCBRewards.append(0)  # 结合未来预测的客户UCB奖励
            self.client_UCB.append(0)  # 客户UCB指标
            self.client_Payment.append(0)  # 客户支付列表(全轮次)
        logging.info("############setup_clients (END)#############")

    def train(self):
        logging.info("self.model_trainer = {}".format(self.model_trainer))
        w_global = self.model_trainer.get_model_params()
        mlops.log_training_status(mlops.ClientConstants.MSG_MLOPS_CLIENT_STATUS_TRAINING)
        mlops.log_aggregation_status(mlops.ServerConstants.MSG_MLOPS_SERVER_STATUS_RUNNING)
        mlops.log_round_info(self.args.comm_round, -1)
        for round_idx in range(self.args.comm_round):
            logging.info("################Communication round : {}".format(round_idx))
            # 本地模型暂存容器
            w_locals = []

            # 客户端选择
            self._client_sampling(round_idx)
            logging.info("client_indexes = " + str(self.client_indexes))

            I_start = time.time()  # 记录本轮开始时间
            for client in enumerate(self.client_indexes):
                # update dataset
                # 判断如果idx是client_indexes中的某一client的下标，那么就更新这个client的数据集
                client.update_local_dataset(
                    client.client_idx,
                    self.train_data_local_dict[client.client_idx],
                    self.test_data_local_dict[client.client_idx],
                    self.train_data_local_num_dict[client.client_idx],
                )
                # train on new dataset
                mlops.event("train", event_started=True,
                            event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
                w_locals.append(client.train(copy.deepcopy(w_global)))
                mlops.event("train", event_started=False,
                            event_value="{}_{}".format(str(round_idx), str(client.client_idx)))
            I_end = time.time()  # 记录本轮结束时间

            # 更细客户完成概率
            self.Completion_Probability(I_start, I_end, round_idx)
            # 根据概率判断是否应该采用该模型
            for i in self.client_indexes:
                if self.client_train_prob[i] == 0: w_locals.remove(i)

            self.CBP(round_idx)  # 计算客户支付

            # update global weights
            mlops.event("agg", event_started=True, event_value=str(round_idx))
            w_global = self._aggregate(w_locals)
            self.model_trainer.set_model_params(w_global)
            mlops.event("agg", event_started=False, event_value=str(round_idx))
            # test results
            # at last round
            train_acc, train_loss, test_acc, test_loss = 0, 0, 0, 0
            if round_idx == self.args.comm_round - 1:
                train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)
            # per {frequency_of_the_test} round
            elif round_idx % self.args.frequency_of_the_test == 0:
                if self.args.dataset.startswith("stackoverflow"):
                    self._local_test_on_validation_set(round_idx)
                else:
                    train_acc, train_loss, test_acc, test_loss = self._local_test_on_all_clients(round_idx)

            mlops.log_round_info(self.args.comm_round, round_idx)

            round_ws.append([round_idx,
                             train_loss,
                             train_acc,
                             time.strftime('%Y-%m-%d %H:%M:%S', time.localtime()),
                             str(self.client_indexes),
                             str(self.client_selected_times)])

            # 保存Excel文件到self.args.excel_save_path+文件名
            wb.save(
                self.args.excel_save_path + self.args.model + "_[" + self.args.dataset + "]_fedavg_training_results_NIID" + str(
                    self.args.experiment_niid_level) + ".xlsx")
            # 休眠一段时间，以便下一个循环开始前有一些时间
            time.sleep(interval)

            self.getReward()  # 更新客户奖励
            self.upgradeUCBParm(round_idx)  # 更新客户UCB参数
            tolPayment = sum(self.client_Payment)
            self.total_budget -= tolPayment  # 更新总预算
            self.upgradedTolReward()  # 更新总奖励
            self.upgradedSel()  # 更新客户被选中次数
            if self.total_budget <= tolPayment:
                # 输出运行时数据：每轮的选择客户、每个客户每轮的支付、剩余预算、每轮每个客户的奖励
                break

        mlops.log_training_finished_status()
        mlops.log_aggregation_finished_status()

    def Completion_Probability(self, Is, Ie, T):  # 计算本轮的完成概率并加入到概率期望中
        ped = Ie - Is  # 计算全局轮的耗时
        for idx, client in enumerate(self.client_list):  # 计算本轮选中客户的完成概率
            x = np.random.normal(Is + ped / 2, 1)  # 根据dI生成耗时随机变量
            # 比较变量和真实耗时，如果小于等于真实耗时，那么完成，否则不完成
            prob = 1 if x <= ped else 0
            # 更新客户历史完成率
            if client.client_idx in self.client_indexes: client.prob = (client.prob * self.client_selected_times[
                idx] + prob) / self.client_selected_times[idx] + 1
            # 更新客户完成概率
            self.client_train_prob[idx] = (self.client_train_prob[idx] * (T - 1) + prob) / T

    def _client_sampling(self, round_idx):  # 基于UCB指标的客户选择,选择UCB指标前K大的客户
        client_num_in_total = self.args.client_num_in_total
        client_num_per_round = self.args.client_num_per_round
        logging.info("client_num_in_total = %s" % str(client_num_in_total))
        logging.info("client_num_per_round = %s" % str(client_num_per_round))
        if round_idx == 1 or client_num_in_total == client_num_per_round:  # 第一轮随机选择
            self.client_indexes = [client_index for client_index in range(client_num_in_total)]
        else:  # 第一轮选中全部客户，其余按投标分数来
            self.getScore()  # 计算客户的得分
            num_clients = min(client_num_per_round, client_num_in_total)
            for idx, client in enumerate(self.client_list):
                # 客户诚实投标,计算其本轮的得分S
                self.client_UCB[idx] = (
                        self.client_scores[idx] * self.client_rewards[idx] / self.client_list[idx].getBid(1)[0])
            sorted_idx = np.argsort(self.client_UCB)  # 按分数排序的客户端下标
            self.client_indexes = sorted_idx[-1 * num_clients:]  # 选出UCB分数最高的K个客户端
        logging.info("当前轮的UCB取胜客户 = %s" % str(self.client_indexes))

    def getScore(self):
        for idx in enumerate(self.client_indexes):
            # 客户诚实投标,计算其本轮的得分S [0,1]
            self.client_scores[idx] = sigmoid(
                sum(i * j for i, j in zip(self.Q_weights, self.client_list[idx].getQuality())))

    def getReward(self):
        for idx in enumerate(self.client_indexes):
            # 计算客户的本轮的奖励r（这里没弄明白系数，先激活归一化）
            self.client_rewards[idx] = sigmoid(self.client_train_prob[idx] * self.client_list[idx].coe * np.sqrt(
                self.client_list[idx].local_sample_number) * pow(self.client_list[idx].loss, 2))

    # 每轮更新客户UCB参数
    def upgradeUCBParm(self, round_idx):
        for idx in enumerate(self.client_list):
            # 更新客户的平均奖励
            if (self.client_rewards[idx] != 0):
                self.client_avgRewards[idx] = (self.client_avgRewards[idx] * self.client_selected_times[idx] +
                                               self.client_rewards[idx]) / (self.client_selected_times[idx] + 1)
            # 更新客户的UCB奖励
            self.client_UCBRewards[idx] = (self.client_avgRewards[idx] * np.sqrt(
                (self.args.client_num_per_round + 1) * log(round_idx, 2)) / self.client_selected_times[idx])

    # 每轮计算客户的UCB指标
    def getUCB(self):
        # 先计算客户的UCB奖励
        for idx, client in enumerate(self.client_list):
            # 根据本轮的UCB奖励、投标、得分计算UCB指标
            self.client_UCB[idx] = (
                    self.client_scores[idx] * self.client_UCBRewards[idx] / self.client_list[idx].getBid(1)[0])

    def CBP(self, round_idx):  # 基于关键投标的支付计算
        if round_idx != 1:  # 第一轮随机选择
            k = self.client_indexes[self.args.client_num_per_round - 1]  # 获取UCB指标最小的客户的下标
            for idx in enumerate(self.client_indexes):  # 计算所有选中客户的支付
                if self.client_train_prob[idx] == 0:  # 如果客户未完成训练，支付为0
                    self.client_Payment[idx] = 0
                else:
                    self.client_Payment[idx] = ((
                        self.client_UCBRewards[idx] * max(self.client_scores) * self.client_list[k].getBid(1) /
                        self.client_UCBRewards[k] * self.client_scores[k], self.args.Cmax))
        else:  # 第一轮选中全部客户
            for idx in enumerate(self.client_indexes):  # 支付固定为Cmax
                if self.client_train_prob[idx] == 0:
                    self.client_Payment[idx] = 0
                else:
                    self.client_Payment[idx] = self.args.Cmax

    def upgradedTolReward(self):  # 更新总奖励
        sum = 0
        for idx, r in self.client_rewards:
            sum += r
            self.total_reward[idx] = 0
        self.total_reward = self.total_reward + sum

    def upgradedSel(self):  # 更新客户被选中次数
        for i in self.client_indexes:
            self.client_selected_times[i] += 1

    def _generate_validation_set(self, num_samples=10000):
        test_data_num = len(self.test_global.dataset)
        sample_indices = random.sample(range(test_data_num), min(num_samples, test_data_num))
        subset = torch.utils.data.Subset(self.test_global.dataset, sample_indices)
        sample_testset = torch.utils.data.DataLoader(subset, batch_size=self.args.batch_size)
        self.val_global = sample_testset

    def _aggregate(self, w_locals):
        training_num = 0
        for idx in range(len(w_locals)):
            (sample_num, averaged_params) = w_locals[idx]
            training_num += sample_num

        (sample_num, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            print("______k = " + str(k))
            for i in range(0, len(w_locals)):
                local_sample_number, local_model_params = w_locals[i]
                w = local_sample_number / training_num
                if i == 0:
                    averaged_params[k] = local_model_params[k] * w
                else:
                    averaged_params[k] += local_model_params[k] * w
        return averaged_params

    def _aggregate_resnet(self, w_locals):  # 弃用
        averaged_params = {}

        clients_tensor = torch.tensor([1.0] * len(w_locals))

        for client in w_locals:
            sample_num, local_params = client

            for key in local_params:
                if key not in averaged_params:
                    averaged_params[key] = torch.zeros_like(local_params[key])

                averaged_params[key] += local_params[key]

        for key in averaged_params:
            averaged_params[key] = averaged_params[key] / clients_tensor

        return averaged_params

    def _aggregate_rnn(self, w_locals):  # 弃用
        # 保存聚合后的参数
        averaged_params = OrderedDict()

        for name, param in w_locals[0][1].named_parameters():

            # 初始化参数均值
            averaged_param = torch.zeros_like(param.data)

            for i in range(len(w_locals)):

                # 获取本地模型的参数
                local_params = w_locals[i][1].named_parameters()

                # 统一使用 1/n 的权重
                w = 1 / len(w_locals)

                # 针对LSTM权重参数做特殊处理
                if 'lstm.weight_hh' in name or 'lstm.weight_ih' in name:
                    averaged_param += local_params[name].data * w.unsqueeze(-1).unsqueeze(-1)

                else:
                    averaged_param += local_params[name].data * w

            # 保存参数均值
            averaged_params[name] = averaged_param

        return averaged_params

    def _aggregate_noniid_avg(self, w_locals):
        """
        The old aggregate method will impact the model performance when it comes to Non-IID setting
        Args:
            w_locals:
        Returns:
        """
        (_, averaged_params) = w_locals[0]
        for k in averaged_params.keys():
            temp_w = []
            for (_, local_w) in w_locals:
                temp_w.append(local_w[k])
            averaged_params[k] = sum(temp_w) / len(temp_w)
        return averaged_params

    def _local_test_on_all_clients(self, round_idx):

        logging.info("################local_test_on_all_clients : {}".format(round_idx))

        train_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        test_metrics = {"num_samples": [], "num_correct": [], "losses": []}

        client = self.client_list[0]

        for client_idx in range(self.args.client_num_in_total):
            """
            Note: for datasets like "fed_CIFAR100" and "fed_shakespheare",
            the training client number is larger than the testing client number
            """
            if self.test_data_local_dict[client_idx] is None:
                continue
            client.update_local_dataset(
                0,
                self.train_data_local_dict[client_idx],
                self.test_data_local_dict[client_idx],
                self.train_data_local_num_dict[client_idx],
            )
            # train data
            train_local_metrics = client.local_test(False)
            train_metrics["num_samples"].append(copy.deepcopy(train_local_metrics["test_total"]))
            train_metrics["num_correct"].append(copy.deepcopy(train_local_metrics["test_correct"]))
            train_metrics["losses"].append(copy.deepcopy(train_local_metrics["test_loss"]))

            # test data
            test_local_metrics = client.local_test(True)
            test_metrics["num_samples"].append(copy.deepcopy(test_local_metrics["test_total"]))
            test_metrics["num_correct"].append(copy.deepcopy(test_local_metrics["test_correct"]))
            test_metrics["losses"].append(copy.deepcopy(test_local_metrics["test_loss"]))

            client_ws.append([round_idx,
                              client_idx,
                              train_metrics["losses"][client_idx] / train_metrics["num_samples"][client_idx],
                              train_metrics["num_correct"][client_idx] / train_metrics["num_samples"][client_idx],
                              time.strftime('%Y-%m-%d %H:%M:%S', time.localtime())])

        # test on training dataset
        train_acc = sum(train_metrics["num_correct"]) / sum(train_metrics["num_samples"])
        train_loss = sum(train_metrics["losses"]) / sum(train_metrics["num_samples"])

        # test on test dataset
        test_acc = sum(test_metrics["num_correct"]) / sum(test_metrics["num_samples"])
        test_loss = sum(test_metrics["losses"]) / sum(test_metrics["num_samples"])

        stats = {"training_acc": train_acc, "training_loss": train_loss}
        if self.args.enable_wandb:
            wandb.log({"Train/Acc": train_acc, "round": round_idx})
            wandb.log({"Train/Loss": train_loss, "round": round_idx})

        # 绘制精度图
        accuracy_list.append(train_acc)
        loss_list.append(train_loss)
        plot_accuracy_and_loss(self, round_idx)

        mlops.log({"Train/Acc": train_acc, "round": round_idx})
        mlops.log({"Train/Loss": train_loss, "round": round_idx})
        logging.info(stats)

        stats = {"test_acc": test_acc, "test_loss": test_loss}
        if self.args.enable_wandb:
            wandb.log({"Test/Acc": test_acc, "round": round_idx})
            wandb.log({"Test/Loss": test_loss, "round": round_idx})

        mlops.log({"Test/Acc": test_acc, "round": round_idx})
        mlops.log({"Test/Loss": test_loss, "round": round_idx})
        logging.info(stats)

        return train_acc, train_loss, test_acc, test_loss

    def _local_test_on_validation_set(self, round_idx):

        logging.info("################local_test_on_validation_set : {}".format(round_idx))

        if self.val_global is None:
            self._generate_validation_set()

        client = self.client_list[0]
        client.update_local_dataset(0, None, self.val_global, None)
        # test data
        test_metrics = client.local_test(True)

        if self.args.dataset == "stackoverflow_nwp":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {"test_acc": test_acc, "test_loss": test_loss}
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})

        elif self.args.dataset == "stackoverflow_lr":
            test_acc = test_metrics["test_correct"] / test_metrics["test_total"]
            test_pre = test_metrics["test_precision"] / test_metrics["test_total"]
            test_rec = test_metrics["test_recall"] / test_metrics["test_total"]
            test_loss = test_metrics["test_loss"] / test_metrics["test_total"]
            stats = {
                "test_acc": test_acc,
                "test_pre": test_pre,
                "test_rec": test_rec,
                "test_loss": test_loss,
            }
            if self.args.enable_wandb:
                wandb.log({"Test/Acc": test_acc, "round": round_idx})
                wandb.log({"Test/Pre": test_pre, "round": round_idx})
                wandb.log({"Test/Rec": test_rec, "round": round_idx})
                wandb.log({"Test/Loss": test_loss, "round": round_idx})

            mlops.log({"Test/Acc": test_acc, "round": round_idx})
            mlops.log({"Test/Pre": test_pre, "round": round_idx})
            mlops.log({"Test/Rec": test_rec, "round": round_idx})
            mlops.log({"Test/Loss": test_loss, "round": round_idx})
        else:
            raise Exception("Unknown format to log metrics for dataset {}!" % self.args.dataset)

        logging.info(stats)


# 定义绘制精度图的函数
def plot_accuracy_and_loss(self, round_idx):
    plt.ion()

    print("accuracy_list: ", accuracy_list)
    print("loss_list: ", loss_list)

    plt.clf()
    plt.suptitle("FedAvg" + "_[" + self.args.dataset + "]_NIID" + str(self.args.experiment_niid_level))

    # 第1个子图
    plt.subplot(1, 3, 1)
    plt.title("accuracy")
    plt.xlabel("num of epoch")
    plt.ylabel("value of accuracy")
    plt.plot(range(1, len(accuracy_list) + 1), accuracy_list, 'b-', linewidth=2)

    # 第2个子图
    plt.subplot(1, 3, 2)
    plt.title("loss")
    plt.xlabel("num of epoch")
    plt.ylabel("value of loss")
    plt.plot(range(1, len(loss_list) + 1), loss_list, 'b-', linewidth=2)

    # 第3个子图，使用条形图展示每个客户的被选择次数
    plt.subplot(1, 3, 3)
    plt.title("num of selected")
    plt.xlabel("num of epoch")
    plt.ylabel("value of num of selected")
    plt.bar(range(1, len(client_selected_times) + 1), client_selected_times, width=0.5, fc='b')

    plt.tight_layout()
    plt.pause(0.005)
    plt.ioff()

    if (round_idx == self.args.comm_round - 1):
        plt.show()

    return
