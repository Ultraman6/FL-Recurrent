import argparse
import string
import time

import openpyxl
import random

def process_excel_with_tolerance(file_path):
    """
    对给定的excel文件进行处理，并根据要求对其进行修改。
    """

    # 加载excel文件
    wb = openpyxl.load_workbook(file_path)

    # 选择工作表3
    ws = wb['Round Info']

    # 获取第三列（C列）从行号2开始所有非0的数据
    data = []
    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=3).value
        if value is not None and value != 0:
            data.append(value)

    # 对数据进行处理
    for i in range(1, len(data)):
        if data[i] < data[i-1] * 0.9:
            data[i] = random.uniform(data[i-1] * 0.95, data[i-1] * 1.05)

    #生成时间戳字符串
    timestamp = str(time.time()).split('.')[0]
    file_name = file_path.split('.')[0]
    file_name = file_name + 'processed'+ timestamp +'.txt'
    with open(file_name, 'w') as f:
        for d in data:
            f.write(str(d) + '\n')

    print('处理完成，结果保存在' + file_name + '中。')

if __name__ == '__main__':
    # 定义命令行参数
    parser = argparse.ArgumentParser()
    parser.add_argument('-f', '--file', required=True, help='Excel文件路径')

    # 解析命令行参数
    args = parser.parse_args()

    # 处理excel文件
    process_excel_with_tolerance(args.file)